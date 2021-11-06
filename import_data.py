from openpyxl import load_workbook
import mysql.connector

# Excel
workbook = load_workbook('imported.xlsx')
sheet = workbook.active

values = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    print(row)
    values.append(row)


# Database
db = mysql.connector.connect(
    host='localhost',
    port=3306,
    user='root',
    password='Admin@123',
    database='Plug_want_to_buy'
)

curses = db.cursor()
sql = '''
    insert into products (title, price, is_necessary)
    values (%s, %s, %s);
'''
curses.executemany(sql, values)
db.commit()
print('เพิ่มข้อมูลจำนวน ' + str(curses.rowcount) + ' แถว')

curses.close()
db.close()