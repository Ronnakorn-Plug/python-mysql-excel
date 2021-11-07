from openpyxl import load_workbook
import mysql.connector

# Excel
workbook = load_workbook('imported2.xlsx')
sheet = workbook.active

# Database
db = mysql.connector.connect(
    host='localhost',
    port=3306,
    user='root',
    password='Admin@123',
    database='Plug_want_to_buy'
)

curses = db.cursor()

# โหลดข้อมูลประเภทสินค้าทั้งหมดออกมาก่อน
sql_sc = '''
    select *
    from categories
'''
curses.execute(sql_sc)
categories = curses.fetchall()


# เปรียบเทียบข้อมูลประเภทสินค้า อันไหนยังไม่มี ให้เพิ่มลงไป
categories_values = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    is_new = True
    category = row[3]

    for c in categories:
        if category == c[1]:
            is_new = False
            break

    if is_new:
        print((category, ))
        categories_values.append((category, ))

if len(categories_values) > 0:
    sql_insert_categories = '''
        insert into categories (title)
        values (%s)
    '''
    curses.executemany(sql_insert_categories, categories_values)
    db.commit()
    print('เพิ่มข้อมูลจำนวน ' + str(curses.rowcount) + ' แถว')
else:
    print('ไม่มีสินค้าใหม่เข้ามาเพิ่ม')

# โหลดประเภทสินค้าทั้งหมด อีกครั้ง
curses.execute(sql_sc)
categories = curses.fetchall()

# เชื่อมต่อ category_id กับสินค้าใหม่จาก excel แล้วเพิ่มลงไป
product_values = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    category_title = row[3]
    category_id = None

    for c in categories:
        if category_title == c[1]:
            category_id = c[0]
            break

    product = (row[0], row[1], row[2], category_id)
    print(product)
    product_values.append(product)

sql_insert_product = '''
    insert into products (title, price, is_necessary, category_id)
    values (%s, %s, %s, %s);
'''
curses.executemany(sql_insert_product, product_values)
db.commit()
print('เพิ่มขสินค้าจำนวน ' + str(curses.rowcount) + ' แถว')


curses.close()
db.close()